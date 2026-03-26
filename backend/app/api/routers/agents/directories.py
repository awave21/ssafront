"""
Роутер для справочников (Directories) агента.

Справочники — структурированные наборы данных, которые автоматически
становятся инструментами (tools) агента.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any
from uuid import UUID

import structlog
from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile, status
from sqlalchemy import delete, func, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, require_scope
from app.api.routers.agents.deps import get_agent_or_404
from app.db.models.directory import Directory, DirectoryItem
from app.db.session import get_db
from app.schemas.auth import AuthContext
from app.schemas.directory import (
    CATALOG_DIRECTORY_TEMPLATES,
    ColumnDefinition,
    DEFAULT_DIRECTORY_TOOL_DESCRIPTION,
    DirectoryCreate,
    DirectoryCreateResponse,
    DirectoryItemCreate,
    DirectoryItemRead,
    DirectoryItemsBulkCreate,
    DirectoryItemsBulkCreateResponse,
    DirectoryItemsBulkDelete,
    DirectoryItemsListResponse,
    DirectoryItemUpdate,
    DirectoryRead,
    DirectorySearchRequest,
    DirectorySearchResponse,
    DirectorySearchResultItem,
    DirectoryToggle,
    DirectoryUpdate,
    ImportPreviewResponse,
    ImportRequest,
    ImportResponse,
    KnowledgeSearchResponse,
    KnowledgeSearchResultItem,
    SINGLETON_DIRECTORY_TEMPLATES,
    STANDARD_QA_DIRECTORY_COLUMNS,
    TEMPLATE_COLUMNS,
)
from app.services.directory.demo_seed import demo_rows_for_template
from app.services.directory.persist_create import (
    DemoItemValidationFailed,
    persist_directory_with_demo_items,
)
from app.services.audit import write_audit

logger = structlog.get_logger(__name__)

router = APIRouter()

# === Лимиты ===
MAX_DIRECTORIES_PER_AGENT = 20
MAX_ITEMS_PER_DIRECTORY = 10_000

TEMPLATE_TOOL_NAME_MAP: dict[str, str] = {
    "qa": "get_question_answer",
    "service_catalog": "get_service_info",
    "product_catalog": "get_product_info",
    "company_info": "get_company_info",
    "theme_catalog": "get_topic_info",
    "medical_course_catalog": "get_medical_course_info",
    "clipboard_import": "get_clipboard_import",
}


def _catalog_tool_description(template: str, when_to_call: str) -> str:
    """Пресет для поля tool_description: имя функции + когда вызывать + параметр query."""
    fn = TEMPLATE_TOOL_NAME_MAP[template]
    return (
        f"Инструмент `{fn}`. {when_to_call} "
        "Параметр `query` — формулировка вопроса пользователя или ключевые слова для поиска по справочнику."
    )


TEMPLATE_TOOL_DESCRIPTION_MAP: dict[str, str] = {
    "qa": _catalog_tool_description(
        "qa",
        "Вызывай, когда нужен короткий точный ответ из базы FAQ (вопрос–ответ).",
    ),
    "service_catalog": _catalog_tool_description(
        "service_catalog",
        "Вызывай для вопросов об услугах: название, описание, стоимость.",
    ),
    "product_catalog": _catalog_tool_description(
        "product_catalog",
        "Вызывай для вопросов о товарах: название, характеристики, цена.",
    ),
    "company_info": _catalog_tool_description(
        "company_info",
        "Вызывай для вопросов о компании: контакты, адрес, доставка, правила, цены.",
    ),
    "theme_catalog": _catalog_tool_description(
        "theme_catalog",
        "Вызывай для вопросов по тематическому каталогу (тема, краткая информация, ссылки).",
    ),
    "medical_course_catalog": _catalog_tool_description(
        "medical_course_catalog",
        "Вызывай для вопросов о медицинских курсах: программа, длительность, стоимость.",
    ),
    "clipboard_import": _catalog_tool_description(
        "clipboard_import",
        "Вызывай для поиска по данным, импортированным из буфера обмена.",
    ),
}


def _catalog_prompt_usage_line(template: str, when_clause: str) -> str:
    """Одна строка для копирования в системный промпт: «Вызывай <fn>, когда …»."""
    fn = TEMPLATE_TOOL_NAME_MAP[template]
    return f"Вызывай {fn}, когда {when_clause}"


# Примеры только для встраивания в промпт агента (не путать с tool_description).
TEMPLATE_PROMPT_USAGE_SNIPPET_MAP: dict[str, str] = {
    "qa": _catalog_prompt_usage_line(
        "qa",
        "пользователю нужен короткий ответ из базы типовых вопросов и ответов (FAQ).",
    ),
    "service_catalog": _catalog_prompt_usage_line(
        "service_catalog",
        "спрашивают об услугах: название, описание, стоимость или подбор услуги.",
    ),
    "product_catalog": _catalog_prompt_usage_line(
        "product_catalog",
        "спрашивают о товарах: название, характеристики, цена или наличие.",
    ),
    "company_info": _catalog_prompt_usage_line(
        "company_info",
        "спрашивают о клинике или компании: контакты, адрес, режим работы, доставка, правила.",
    ),
    "theme_catalog": _catalog_prompt_usage_line(
        "theme_catalog",
        "нужна информация по тематическому каталогу: тема, краткое содержание, ссылки.",
    ),
    "medical_course_catalog": _catalog_prompt_usage_line(
        "medical_course_catalog",
        "спрашивают о медицинских курсах: программа, длительность, стоимость, запись.",
    ),
    "clipboard_import": _catalog_prompt_usage_line(
        "clipboard_import",
        "нужно найти данные из справочника, импортированного из буфера обмена или файла.",
    ),
}


def _default_prompt_usage_snippet_for_custom(tool_name: str) -> str:
    return (
        f"Вызывай {tool_name}, когда нужна информация из этого справочника "
        "(уточни в промпте, по каким темам к нему обращаться)."
    )


def _resolve_prompt_usage_snippet(
    payload: str | None,
    *,
    template: str,
    tool_name: str,
) -> str | None:
    text = (payload or "").strip()
    if text:
        return text
    preset = TEMPLATE_PROMPT_USAGE_SNIPPET_MAP.get(template)
    if preset:
        return preset
    return _default_prompt_usage_snippet_for_custom(tool_name)


def _qa_columns() -> list[dict[str, Any]]:
    return [dict(col) for col in STANDARD_QA_DIRECTORY_COLUMNS]


def _normalize_columns_data(columns: list[ColumnDefinition] | None, template: str) -> list[dict[str, Any]]:
    if template in CATALOG_DIRECTORY_TEMPLATES:
        return _qa_columns()
    if columns is None:
        raise ValueError("columns are required")
    return [col.model_dump() for col in columns]


def _pg_enum_to_str(value: Any) -> str:
    """Строка для полей SQLAlchemy Enum / драйвера — в Pydantic уходит Literal."""
    if isinstance(value, str):
        return value
    if value is not None:
        inner = getattr(value, "value", None)
        if isinstance(inner, str):
            return inner
    return str(value) if value is not None else ""


def _build_directory_create_response(
    directory: Directory,
    columns_data: list[dict[str, Any]],
    *,
    seeded_demo_items_count: int,
) -> DirectoryCreateResponse:
    """
    Ответ POST собираем из явных полей и columns_data (источник до round-trip JSONB).
    Так клиент не получает 500 после успешного commit из‑за model_validate по ORM.
    """
    columns = [ColumnDefinition.model_validate(c) for c in columns_data]
    template = _pg_enum_to_str(directory.template) or "custom"
    response_mode = _pg_enum_to_str(directory.response_mode) or "function_result"
    search_type = _pg_enum_to_str(directory.search_type) or "fuzzy"
    return DirectoryCreateResponse.model_validate(
        {
            "id": directory.id,
            "tenant_id": directory.tenant_id,
            "agent_id": directory.agent_id,
            "name": directory.name,
            "slug": directory.slug,
            "tool_name": directory.tool_name,
            "tool_description": directory.tool_description,
            "prompt_usage_snippet": directory.prompt_usage_snippet,
            "template": template,
            "columns": columns,
            "response_mode": response_mode,
            "search_type": search_type,
            "is_enabled": directory.is_enabled,
            "items_count": int(directory.items_count or 0),
            "created_at": directory.created_at,
            "updated_at": directory.updated_at,
            "seeded_demo_items_count": seeded_demo_items_count,
        }
    )


def _resolve_tool_name(payload_tool_name: str | None, *, template: str, name: str) -> str:
    fixed = TEMPLATE_TOOL_NAME_MAP.get(template)
    if fixed:
        return fixed
    if payload_tool_name:
        return payload_tool_name
    # Fallback for custom templates: keep deterministic generation from directory name.
    return _generate_slug(name).replace("-", "_")


def _resolve_tool_description(
    payload_description: str | None,
    *,
    template: str,
    tool_name: str | None = None,
) -> str:
    text = (payload_description or "").strip()
    if text:
        return text
    preset = TEMPLATE_TOOL_DESCRIPTION_MAP.get(template)
    if preset:
        return preset
    if tool_name:
        return (
            f"Инструмент `{tool_name}`. Поиск по этому справочнику. "
            "Параметр `query` — формулировка вопроса пользователя или ключевые слова."
        )
    logger.info("directory_tool_description_fallback", template=template)
    return DEFAULT_DIRECTORY_TOOL_DESCRIPTION


def _generate_slug(name: str) -> str:
    """Генерирует slug из названия справочника."""
    # Транслитерация кириллицы
    translit_map = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    result = name.lower()
    for cyr, lat in translit_map.items():
        result = result.replace(cyr, lat)
    # Заменяем всё кроме букв и цифр на дефис
    result = re.sub(r"[^a-z0-9]+", "-", result)
    # Убираем начальные и конечные дефисы
    result = result.strip("-")
    return result or "directory"


async def _get_directory_or_404(
    directory_id: UUID,
    agent_id: UUID,
    db: AsyncSession,
    user: AuthContext,
) -> Directory:
    """Получить справочник или вернуть 404."""
    stmt = select(Directory).where(
        Directory.id == directory_id,
        Directory.agent_id == agent_id,
        Directory.tenant_id == user.tenant_id,
        Directory.is_deleted.is_(False),
    )
    directory = (await db.execute(stmt)).scalar_one_or_none()
    if directory is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Directory not found",
        )
    return directory


def _validate_item_data(
    data: dict[str, Any],
    columns: list[dict],
    row_num: int | None = None,
) -> list[str]:
    """
    Валидирует данные записи согласно определению колонок.
    
    Поддерживаемые типы: text, number, date, bool
    
    Возвращает список ошибок. Пустой список = валидация успешна.
    """
    errors: list[str] = []
    prefix = f"Row {row_num}: " if row_num is not None else ""
    
    for col in columns:
        col_name = col["name"]
        col_type = col["type"]
        required = col.get("required", False)
        value = data.get(col_name)
        
        # Проверка обязательности
        if required and (value is None or value == ""):
            errors.append(f"{prefix}Field '{col_name}' is required")
            continue
        
        if value is None or value == "":
            continue
        
        # Проверка типов (4 типа: text, number, date, bool)
        try:
            if col_type == "text":
                # Любая строка
                if not isinstance(value, str):
                    # Попробуем конвертировать
                    data[col_name] = str(value)
            
            elif col_type == "number":
                # Число (целое или с точкой)
                if isinstance(value, bool):
                    errors.append(f"{prefix}Field '{col_name}' must be a number, not boolean")
                elif isinstance(value, (int, float)):
                    pass  # OK
                elif isinstance(value, str):
                    # Попробуем конвертировать строку в число
                    try:
                        data[col_name] = float(value) if "." in value else int(value)
                    except ValueError:
                        errors.append(f"{prefix}Field '{col_name}' must be a number")
                else:
                    errors.append(f"{prefix}Field '{col_name}' must be a number")
            
            elif col_type == "date":
                # Формат YYYY-MM-DD
                if isinstance(value, str):
                    if not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
                        errors.append(f"{prefix}Field '{col_name}' must be date in YYYY-MM-DD format")
                    else:
                        # Проверим, что дата валидная
                        try:
                            datetime.strptime(value, "%Y-%m-%d")
                        except ValueError:
                            errors.append(f"{prefix}Field '{col_name}' is not a valid date")
                else:
                    errors.append(f"{prefix}Field '{col_name}' must be a date string (YYYY-MM-DD)")
            
            elif col_type == "bool":
                # true / false
                if not isinstance(value, bool):
                    # Пробуем конвертировать
                    if isinstance(value, str):
                        if value.lower() in ("true", "1", "yes"):
                            data[col_name] = True
                        elif value.lower() in ("false", "0", "no"):
                            data[col_name] = False
                        else:
                            errors.append(f"{prefix}Field '{col_name}' must be a boolean")
                    elif isinstance(value, int):
                        data[col_name] = bool(value)
                    else:
                        errors.append(f"{prefix}Field '{col_name}' must be a boolean")
            
            # Legacy типы (для обратной совместимости)
            elif col_type in ("varchar", "numeric", "integer", "bigint", "boolean"):
                # Обрабатываем как алиасы
                if col_type == "varchar":
                    if not isinstance(value, str):
                        data[col_name] = str(value)
                elif col_type in ("numeric", "integer", "bigint"):
                    if not isinstance(value, (int, float)) or isinstance(value, bool):
                        try:
                            data[col_name] = float(value)
                        except (ValueError, TypeError):
                            errors.append(f"{prefix}Field '{col_name}' must be a number")
                elif col_type == "boolean":
                    if not isinstance(value, bool):
                        errors.append(f"{prefix}Field '{col_name}' must be a boolean")
        
        except (ValueError, TypeError) as e:
            errors.append(f"{prefix}Field '{col_name}' validation error: {e}")
    
    return errors


# === Directory CRUD ===


@router.get("", response_model=list[DirectoryRead])
async def list_directories(
    agent_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> list[DirectoryRead]:
    """Получить список справочников агента."""
    # Проверяем, что агент существует и принадлежит пользователю
    await get_agent_or_404(agent_id, db, user)
    
    stmt = (
        select(Directory)
        .where(
            Directory.agent_id == agent_id,
            Directory.tenant_id == user.tenant_id,
            Directory.is_deleted.is_(False),
        )
        .order_by(Directory.created_at.desc())
    )
    result = await db.execute(stmt)
    directories = result.scalars().all()
    return [DirectoryRead.model_validate(d) for d in directories]


@router.post("", response_model=DirectoryCreateResponse, status_code=status.HTTP_201_CREATED)
async def create_directory(
    agent_id: UUID,
    payload: DirectoryCreate,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryCreateResponse:
    """Создать новый справочник (опционально 3 демо-записи в той же транзакции)."""
    # Проверяем, что агент существует
    await get_agent_or_404(agent_id, db, user)
    
    # Проверяем лимит справочников
    count_stmt = select(func.count(Directory.id)).where(
        Directory.agent_id == agent_id,
        Directory.tenant_id == user.tenant_id,
        Directory.is_deleted.is_(False),
    )
    count = (await db.execute(count_stmt)).scalar() or 0
    if count >= MAX_DIRECTORIES_PER_AGENT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Maximum {MAX_DIRECTORIES_PER_AGENT} directories per agent allowed",
        )

    if payload.template in SINGLETON_DIRECTORY_TEMPLATES:
        dup_stmt = (
            select(Directory.id)
            .where(
                Directory.agent_id == agent_id,
                Directory.tenant_id == user.tenant_id,
                Directory.template == payload.template,
                Directory.is_deleted.is_(False),
            )
            .limit(1)
        )
        if (await db.execute(dup_stmt)).scalar_one_or_none() is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A directory with this template already exists for this agent",
            )

    # Генерируем slug
    slug = _generate_slug(payload.name)
    
    # Подготавливаем колонки
    columns_data = _normalize_columns_data(payload.columns, payload.template)
    search_type = (
        "semantic" if payload.template in CATALOG_DIRECTORY_TEMPLATES else payload.search_type
    )

    resolved_tool_name = _resolve_tool_name(payload.tool_name, template=payload.template, name=payload.name)
    resolved_tool_description = _resolve_tool_description(
        payload.tool_description,
        template=payload.template,
        tool_name=resolved_tool_name,
    )
    resolved_prompt_snippet = _resolve_prompt_usage_snippet(
        payload.prompt_usage_snippet,
        template=payload.template,
        tool_name=resolved_tool_name,
    )
    directory = Directory(
        tenant_id=user.tenant_id,
        agent_id=agent_id,
        name=payload.name,
        slug=slug,
        tool_name=resolved_tool_name,
        tool_description=resolved_tool_description,
        prompt_usage_snippet=resolved_prompt_snippet,
        template=payload.template,
        columns=columns_data,
        response_mode=payload.response_mode,
        search_type=search_type,
    )

    demo_rows = demo_rows_for_template(payload.template) if payload.seed_demo_items else []
    seeded_demo_items_count = 0

    try:
        await persist_directory_with_demo_items(
            db,
            directory=directory,
            columns_data=columns_data,
            demo_rows=demo_rows,
            tenant_id=user.tenant_id,
            validate_item_row=lambda d, cols, rn: _validate_item_data(d, cols, row_num=rn),
        )
        seeded_demo_items_count = len(demo_rows)
    except DemoItemValidationFailed as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=exc.errors,
        ) from exc
    except IntegrityError as exc:
        await db.rollback()
        if "uq_directories_agent_slug" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Directory with this slug already exists",
            ) from exc
        if "uq_directories_agent_tool_name" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Directory with this tool_name already exists",
            ) from exc
        if "uq_directories_agent_singleton_template" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A directory with this template already exists for this agent",
            ) from exc
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Directory already exists",
        ) from exc

    await db.refresh(directory)

    # Каталожные справочники — semantic: без эмбеддингов поиск по вектору пустой (embedding IS NULL).
    if _pg_enum_to_str(directory.search_type) == "semantic" and directory.items_count > 0:
        from app.services.directory.service import update_directory_embeddings

        await update_directory_embeddings(db, directory)

    response = _build_directory_create_response(
        directory,
        columns_data,
        seeded_demo_items_count=seeded_demo_items_count,
    )

    try:
        await write_audit(db, user, "directory.create", "directory", str(directory.id))
    except Exception as exc:
        # Справочник уже закоммичен в persist_directory_with_demo_items; аудит не должен
        # превращать успешное создание в 500 для клиента.
        logger.error(
            "directory_create_audit_failed",
            directory_id=str(directory.id),
            agent_id=str(agent_id),
            error=str(exc),
            exc_info=True,
        )

    logger.info(
        "directory_created",
        directory_id=str(directory.id),
        agent_id=str(agent_id),
        template=payload.template,
        seeded_demo_items_count=seeded_demo_items_count,
    )

    return response


@router.get("/{directory_id}", response_model=DirectoryRead)
async def get_directory(
    agent_id: UUID,
    directory_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryRead:
    """Получить справочник по ID."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    return DirectoryRead.model_validate(directory)


@router.put("/{directory_id}", response_model=DirectoryRead)
async def update_directory(
    agent_id: UUID,
    directory_id: UUID,
    payload: DirectoryUpdate,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryRead:
    """Обновить справочник."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    update_data = payload.model_dump(exclude_unset=True)
    
    # Запоминаем, меняется ли search_type на semantic
    old_search_type = directory.search_type
    new_search_type = update_data.get("search_type", old_search_type)
    switching_to_semantic = (old_search_type != "semantic" and new_search_type == "semantic")
    
    # Если меняется имя — обновляем slug
    if "name" in update_data:
        update_data["slug"] = _generate_slug(update_data["name"])

    tpl_key = _pg_enum_to_str(directory.template) or "custom"
    fixed_tool_name = TEMPLATE_TOOL_NAME_MAP.get(tpl_key)
    if fixed_tool_name:
        # Template-based directories keep stable function name.
        update_data["tool_name"] = fixed_tool_name
        if "tool_description" in update_data and not str(update_data["tool_description"] or "").strip():
            update_data["tool_description"] = (
                TEMPLATE_TOOL_DESCRIPTION_MAP.get(tpl_key)
                or DEFAULT_DIRECTORY_TOOL_DESCRIPTION
            )
    elif "tool_name" in update_data and not update_data["tool_name"]:
        update_data["tool_name"] = _generate_slug(update_data.get("name", directory.name)).replace("-", "_")

    if "prompt_usage_snippet" in update_data and not str(
        update_data["prompt_usage_snippet"] or ""
    ).strip():
        tool_for_snippet = fixed_tool_name or update_data.get("tool_name", directory.tool_name)
        update_data["prompt_usage_snippet"] = (
            TEMPLATE_PROMPT_USAGE_SNIPPET_MAP.get(tpl_key)
            or _default_prompt_usage_snippet_for_custom(str(tool_for_snippet))
        )
    
    if directory.template != "custom" and "columns" in update_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="columns are fixed for catalog directories",
        )
    if (
        directory.template != "custom"
        and "search_type" in update_data
        and update_data["search_type"] != "semantic"
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="search_type for catalog directories is fixed to semantic",
        )

    # Если меняются колонки — нужно мигрировать данные
    if "columns" in update_data and update_data["columns"] is not None:
        old_columns = {col["name"]: col for col in directory.columns}
        # columns уже являются dict после model_dump()
        new_columns = {col["name"]: col for col in update_data["columns"]}
        
        # Удалённые колонки
        removed_cols = set(old_columns.keys()) - set(new_columns.keys())
        
        # Удаляем поля из всех записей
        if removed_cols:
            for col_name in removed_cols:
                await db.execute(
                    text(
                        f"""
                        UPDATE directory_items 
                        SET data = data - '{col_name}', updated_at = NOW()
                        WHERE directory_id = :directory_id
                        """
                    ),
                    {"directory_id": str(directory_id)},
                )
        
        # columns уже в формате list[dict], не нужно конвертировать
    
    for key, value in update_data.items():
        setattr(directory, key, value)
    
    try:
        await db.commit()
    except IntegrityError as exc:
        await db.rollback()
        if "uq_directories_agent_slug" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Directory with this slug already exists",
            ) from exc
        if "uq_directories_agent_tool_name" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Directory with this tool_name already exists",
            ) from exc
        if "uq_directories_agent_singleton_template" in str(exc):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A directory with this template already exists for this agent",
            ) from exc
        raise
    
    await db.refresh(directory)
    
    # Если search_type = semantic — генерируем недостающие embeddings
    if directory.search_type == "semantic" and directory.items_count > 0:
        from app.services.directory.service import update_directory_embeddings
        updated = await update_directory_embeddings(db, directory)
        if updated > 0:
            logger.info(
                "embeddings_generated",
                directory_id=str(directory.id),
                updated=updated,
            )
    
    await write_audit(db, user, "directory.update", "directory", str(directory.id))
    
    return DirectoryRead.model_validate(directory)


@router.delete("/{directory_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_directory(
    agent_id: UUID,
    directory_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> None:
    """Удалить справочник (hard delete)."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    directory_id_str = str(directory.id)
    
    # Hard delete — удаляем из базы полностью
    # Записи DirectoryItem удалятся автоматически (cascade)
    await db.delete(directory)
    await db.commit()
    await write_audit(db, user, "directory.delete", "directory", directory_id_str)
    
    logger.info(
        "directory_deleted",
        directory_id=directory_id_str,
        agent_id=str(agent_id),
    )


@router.patch("/{directory_id}/toggle", response_model=DirectoryRead)
async def toggle_directory(
    agent_id: UUID,
    directory_id: UUID,
    payload: DirectoryToggle,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryRead:
    """Быстрое включение/выключение справочника."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    directory.is_enabled = payload.is_enabled
    await db.commit()
    await db.refresh(directory)
    await write_audit(db, user, "directory.toggle", "directory", str(directory.id))
    
    return DirectoryRead.model_validate(directory)


# === Directory Items CRUD ===


@router.get("/{directory_id}/items", response_model=DirectoryItemsListResponse)
async def list_directory_items(
    agent_id: UUID,
    directory_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
    limit: int = Query(default=50, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    search: str | None = Query(default=None, max_length=200),
) -> DirectoryItemsListResponse:
    """Получить записи справочника с пагинацией и поиском."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Базовый запрос
    base_where = [
        DirectoryItem.directory_id == directory_id,
        DirectoryItem.tenant_id == user.tenant_id,
    ]
    
    # Если есть поиск — добавляем условие
    if search:
        # Ищем по searchable колонкам через ILIKE
        searchable_cols = [col["name"] for col in directory.columns if col.get("searchable")]
        if searchable_cols:
            search_conditions = " OR ".join(
                f"data->>'{col}' ILIKE :search" for col in searchable_cols
            )
            # Используем raw SQL для поиска
            items_stmt = f"""
                SELECT * FROM directory_items
                WHERE directory_id = :directory_id 
                  AND tenant_id = :tenant_id
                  AND ({search_conditions})
                ORDER BY created_at DESC
                LIMIT :limit OFFSET :offset
            """
            count_stmt = f"""
                SELECT COUNT(*) FROM directory_items
                WHERE directory_id = :directory_id 
                  AND tenant_id = :tenant_id
                  AND ({search_conditions})
            """
            params = {
                "directory_id": directory_id,
                "tenant_id": user.tenant_id,
                "search": f"%{search}%",
                "limit": limit,
                "offset": offset,
            }
            
            result = await db.execute(items_stmt, params)  # type: ignore
            items_raw = result.fetchall()
            
            count_result = await db.execute(count_stmt, params)  # type: ignore
            total = count_result.scalar() or 0
            
            # Конвертируем в модели
            items = [
                DirectoryItemRead(
                    id=row.id,
                    directory_id=row.directory_id,
                    data=row.data,
                    created_at=row.created_at,
                    updated_at=row.updated_at,
                )
                for row in items_raw
            ]
        else:
            # Нет searchable колонок — возвращаем пустой результат
            items = []
            total = 0
    else:
        # Без поиска — обычный SELECT
        stmt = (
            select(DirectoryItem)
            .where(*base_where)
            .order_by(DirectoryItem.created_at.desc())
            .limit(limit)
            .offset(offset)
        )
        result = await db.execute(stmt)
        items_db = result.scalars().all()
        items = [DirectoryItemRead.model_validate(item) for item in items_db]
        
        count_stmt = select(func.count(DirectoryItem.id)).where(*base_where)
        total = (await db.execute(count_stmt)).scalar() or 0
    
    return DirectoryItemsListResponse(
        items=items,
        total=total,
        limit=limit,
        offset=offset,
    )


@router.post("/{directory_id}/items", response_model=DirectoryItemRead, status_code=status.HTTP_201_CREATED)
async def create_directory_item(
    agent_id: UUID,
    directory_id: UUID,
    payload: DirectoryItemCreate,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryItemRead:
    """Добавить одну запись в справочник."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Проверяем лимит записей
    if directory.items_count >= MAX_ITEMS_PER_DIRECTORY:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Maximum {MAX_ITEMS_PER_DIRECTORY} items per directory allowed",
        )
    
    # Валидируем данные
    errors = _validate_item_data(payload.data, directory.columns)
    if errors:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=errors,
        )
    
    item = DirectoryItem(
        tenant_id=user.tenant_id,
        directory_id=directory_id,
        data=payload.data,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    
    # Триггер автоматически обновит items_count

    if _pg_enum_to_str(directory.search_type) == "semantic":
        from app.services.directory.service import update_item_embedding

        await update_item_embedding(db, item, directory)
    
    return DirectoryItemRead.model_validate(item)


@router.post("/{directory_id}/items/bulk", response_model=DirectoryItemsBulkCreateResponse)
async def bulk_create_directory_items(
    agent_id: UUID,
    directory_id: UUID,
    payload: DirectoryItemsBulkCreate,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryItemsBulkCreateResponse:
    """Массовое добавление записей."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Если replace_all — удаляем все существующие записи
    if payload.replace_all:
        await db.execute(
            delete(DirectoryItem).where(
                DirectoryItem.directory_id == directory_id,
                DirectoryItem.tenant_id == user.tenant_id,
            )
        )
        # Сбрасываем счётчик (триггер не сработает при DELETE с WHERE)
        directory.items_count = 0
    
    # Проверяем лимит
    remaining = MAX_ITEMS_PER_DIRECTORY - directory.items_count
    if len(payload.items) > remaining:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot add {len(payload.items)} items. Maximum {remaining} items can be added.",
        )
    
    created = 0
    errors: list[dict[str, Any]] = []
    
    for i, item_data in enumerate(payload.items, start=1):
        validation_errors = _validate_item_data(item_data.data, directory.columns, row_num=i)
        if validation_errors:
            errors.append({"row": i, "error": "; ".join(validation_errors)})
            continue
        
        item = DirectoryItem(
            tenant_id=user.tenant_id,
            directory_id=directory_id,
            data=item_data.data,
        )
        db.add(item)
        created += 1
    
    await db.commit()
    await db.refresh(directory)
    
    logger.info(
        "bulk_items_created",
        directory_id=str(directory_id),
        created=created,
        errors=len(errors),
    )

    if created > 0 and _pg_enum_to_str(directory.search_type) == "semantic":
        from app.services.directory.service import update_directory_embeddings

        await update_directory_embeddings(db, directory)
    
    return DirectoryItemsBulkCreateResponse(created=created, errors=errors)


@router.put("/{directory_id}/items/{item_id}", response_model=DirectoryItemRead)
async def update_directory_item(
    agent_id: UUID,
    directory_id: UUID,
    item_id: UUID,
    payload: DirectoryItemUpdate,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectoryItemRead:
    """Обновить запись справочника."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Получаем запись
    stmt = select(DirectoryItem).where(
        DirectoryItem.id == item_id,
        DirectoryItem.directory_id == directory_id,
        DirectoryItem.tenant_id == user.tenant_id,
    )
    item = (await db.execute(stmt)).scalar_one_or_none()
    if item is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Directory item not found",
        )
    
    # Валидируем данные
    errors = _validate_item_data(payload.data, directory.columns)
    if errors:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=errors,
        )
    
    item.data = payload.data
    await db.commit()
    await db.refresh(item)

    if _pg_enum_to_str(directory.search_type) == "semantic":
        from app.services.directory.service import update_item_embedding

        await update_item_embedding(db, item, directory)
    
    return DirectoryItemRead.model_validate(item)


@router.delete("/{directory_id}/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_directory_item(
    agent_id: UUID,
    directory_id: UUID,
    item_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> None:
    """Удалить одну запись справочника."""
    await get_agent_or_404(agent_id, db, user)
    await _get_directory_or_404(directory_id, agent_id, db, user)
    
    stmt = select(DirectoryItem).where(
        DirectoryItem.id == item_id,
        DirectoryItem.directory_id == directory_id,
        DirectoryItem.tenant_id == user.tenant_id,
    )
    item = (await db.execute(stmt)).scalar_one_or_none()
    if item is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Directory item not found",
        )
    
    await db.delete(item)
    await db.commit()
    # Триггер автоматически уменьшит items_count


@router.delete("/{directory_id}/items", status_code=status.HTTP_204_NO_CONTENT)
async def bulk_delete_directory_items(
    agent_id: UUID,
    directory_id: UUID,
    payload: DirectoryItemsBulkDelete,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> None:
    """Массовое удаление записей."""
    await get_agent_or_404(agent_id, db, user)
    await _get_directory_or_404(directory_id, agent_id, db, user)
    
    await db.execute(
        delete(DirectoryItem).where(
            DirectoryItem.id.in_(payload.ids),
            DirectoryItem.directory_id == directory_id,
            DirectoryItem.tenant_id == user.tenant_id,
        )
    )
    await db.commit()


@router.post("/{directory_id}/rebuild-embeddings")
async def rebuild_directory_embeddings(
    agent_id: UUID,
    directory_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> dict:
    """
    Пересоздать embeddings для всех записей справочника.
    
    Используется для справочников с search_type='semantic'.
    Операция может занять время при большом количестве записей.
    """
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    if directory.search_type != "semantic":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Directory search_type must be 'semantic' to use embeddings",
        )
    
    from app.services.directory.service import update_directory_embeddings
    
    updated = await update_directory_embeddings(db, directory)
    
    logger.info(
        "embeddings_rebuilt",
        directory_id=str(directory_id),
        updated=updated,
    )
    
    return {"updated": updated, "message": f"Successfully updated {updated} embeddings"}


# === Search ===


@router.post("/{directory_id}/search", response_model=DirectorySearchResponse)
async def search_directory(
    agent_id: UUID,
    directory_id: UUID,
    payload: DirectorySearchRequest,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> DirectorySearchResponse:
    """Поиск по справочнику."""
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Импортируем сервис поиска
    from app.services.directory.service import search_directory_items
    
    results = await search_directory_items(
        db=db,
        directory=directory,
        query=payload.query,
        limit=payload.limit,
    )
    
    return DirectorySearchResponse(results=results)


@router.post("/knowledge/search", response_model=KnowledgeSearchResponse)
async def search_all_directories(
    agent_id: UUID,
    payload: DirectorySearchRequest,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> KnowledgeSearchResponse:
    """Поиск по ВСЕМ справочникам агента."""
    await get_agent_or_404(agent_id, db, user)
    
    # Получаем все активные справочники агента
    stmt = select(Directory).where(
        Directory.agent_id == agent_id,
        Directory.tenant_id == user.tenant_id,
        Directory.is_deleted.is_(False),
        Directory.is_enabled.is_(True),
    )
    directories = (await db.execute(stmt)).scalars().all()
    
    if not directories:
        return KnowledgeSearchResponse(results=[])
    
    from app.services.directory.service import search_directory_items
    
    all_results: list[KnowledgeSearchResultItem] = []
    
    for directory in directories:
        results = await search_directory_items(
            db=db,
            directory=directory,
            query=payload.query,
            limit=payload.limit,
        )
        
        for item in results:
            all_results.append(
                KnowledgeSearchResultItem(
                    directory_id=directory.id,
                    directory_name=directory.name,
                    item=item,
                    relevance=item.relevance,
                )
            )
    
    # Сортируем по релевантности и ограничиваем
    all_results.sort(key=lambda x: x.relevance, reverse=True)
    
    return KnowledgeSearchResponse(results=all_results[: payload.limit])


# === Import/Export ===


@router.post("/{directory_id}/import/preview", response_model=ImportPreviewResponse)
async def preview_import(
    agent_id: UUID,
    directory_id: UUID,
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> ImportPreviewResponse:
    """
    Превью импорта — парсинг файла без сохранения.
    
    Возвращает колонки файла, количество строк и предложенный маппинг.
    """
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Импортируем сервис импорта
    from app.services.directory.importer import parse_file_preview, suggest_column_mapping
    
    # Парсим файл
    columns, rows_count, preview_rows = await parse_file_preview(file)
    
    # Предлагаем маппинг
    suggested_mapping = suggest_column_mapping(columns, directory.columns)
    
    return ImportPreviewResponse(
        columns=columns,
        rows_count=rows_count,
        preview=preview_rows,
        suggested_mapping=suggested_mapping,
    )


@router.post("/{directory_id}/import", response_model=ImportResponse)
async def import_directory(
    agent_id: UUID,
    directory_id: UUID,
    file: UploadFile,
    mapping: str = Form(..., description="JSON маппинг колонок файла → колонки справочника"),
    has_header: bool = Form(default=True),
    replace_all: bool = Form(default=False),
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
) -> ImportResponse:
    """
    Импорт данных из CSV/Excel файла.
    
    Параметры:
    - file: CSV или XLSX файл
    - mapping: JSON маппинг колонок файла → колонки справочника
    - has_header: есть ли заголовок в файле
    - replace_all: заменить все существующие данные
    """
    import json as json_module
    
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Парсим маппинг
    try:
        mapping_dict = json_module.loads(mapping)
    except json_module.JSONDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid mapping JSON",
        )
    
    # Импортируем сервис
    from app.services.directory.importer import import_file_data
    
    # Проверяем лимит
    if replace_all:
        # Удаляем все записи
        await db.execute(
            delete(DirectoryItem).where(
                DirectoryItem.directory_id == directory_id,
                DirectoryItem.tenant_id == user.tenant_id,
            )
        )
        directory.items_count = 0
    
    remaining = MAX_ITEMS_PER_DIRECTORY - directory.items_count
    
    # Импортируем данные
    created, skipped, errors = await import_file_data(
        db=db,
        file=file,
        mapping=mapping_dict,
        has_header=has_header,
        directory=directory,
        tenant_id=user.tenant_id,
        max_items=remaining,
        validate_fn=lambda data, row_num: _validate_item_data(data, directory.columns, row_num),
    )
    
    await db.commit()
    await db.refresh(directory)
    
    logger.info(
        "import_completed",
        directory_id=str(directory_id),
        created=created,
        skipped=skipped,
        errors=len(errors),
    )

    if created > 0 and _pg_enum_to_str(directory.search_type) == "semantic":
        from app.services.directory.service import update_directory_embeddings

        await update_directory_embeddings(db, directory)
    
    return ImportResponse(created=created, skipped=skipped, errors=errors)


@router.get("/{directory_id}/export")
async def export_directory(
    agent_id: UUID,
    directory_id: UUID,
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
    user: AuthContext = Depends(require_scope("agents:write")),
):
    """
    Экспорт справочника в CSV или XLSX.
    
    Возвращает файл для скачивания.
    
    Авторизация: Authorization: Bearer <token>
    """
    from fastapi.responses import StreamingResponse
    import io
    
    await get_agent_or_404(agent_id, db, user)
    directory = await _get_directory_or_404(directory_id, agent_id, db, user)
    
    # Получаем все записи
    stmt = (
        select(DirectoryItem)
        .where(
            DirectoryItem.directory_id == directory_id,
            DirectoryItem.tenant_id == user.tenant_id,
        )
        .order_by(DirectoryItem.created_at.desc())
    )
    result = await db.execute(stmt)
    items = result.scalars().all()
    
    # Формируем данные
    columns = [col["name"] for col in directory.columns]
    labels = [col["label"] for col in directory.columns]
    
    if format == "csv":
        import csv
        
        output = io.StringIO()
        # Используем точку с запятой как разделитель для лучшей совместимости
        # с Excel на русской локали
        writer = csv.writer(output, delimiter=';')
        
        # Заголовок
        writer.writerow(labels)
        
        # Данные
        for item in items:
            row = [item.data.get(col, "") for col in columns]
            writer.writerow(row)
        
        output.seek(0)
        csv_content = output.getvalue()
        
        # Добавляем BOM для UTF-8, чтобы Excel правильно распознал кодировку
        csv_bytes = b'\xef\xbb\xbf' + csv_content.encode('utf-8')
        
        return StreamingResponse(
            iter([csv_bytes]),
            media_type="text/csv; charset=utf-8-sig",
            headers={
                "Content-Disposition": f'attachment; filename="{directory.slug}.csv"',
            },
        )
    
    else:  # xlsx
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="XLSX export requires openpyxl library",
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = directory.name[:31]  # Max 31 chars for sheet name
        
        # Заголовок
        for col_idx, label in enumerate(labels, start=1):
            ws.cell(row=1, column=col_idx, value=label)
        
        # Данные
        for row_idx, item in enumerate(items, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=item.data.get(col_name, ""))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{directory.slug}.xlsx"',
            },
        )
