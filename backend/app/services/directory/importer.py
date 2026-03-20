"""
Сервис импорта данных для справочников.

Поддерживает импорт из CSV и Excel файлов.
"""

from __future__ import annotations

import csv
import io
from typing import Any, Callable

import structlog
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.directory import Directory, DirectoryItem

logger = structlog.get_logger(__name__)

# Максимальный размер файла (10 МБ)
MAX_FILE_SIZE = 10 * 1024 * 1024

# Максимальное количество строк для превью
PREVIEW_ROWS = 5


async def parse_file_preview(
    file: Any,
) -> tuple[list[str], int, list[list[Any]]]:
    """
    Парсит файл и возвращает превью.
    
    Args:
        file: UploadFile объект
    
    Returns:
        Tuple из:
        - columns: список названий колонок
        - rows_count: общее количество строк
        - preview: первые N строк данных
    """
    # Читаем содержимое файла
    content = await file.read()
    await file.seek(0)  # Сбрасываем позицию для повторного чтения
    
    if len(content) > MAX_FILE_SIZE:
        raise ValueError(f"File too large. Maximum size is {MAX_FILE_SIZE // 1024 // 1024} MB")
    
    filename = getattr(file, "filename", "") or ""
    
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return _parse_excel_preview(content)
    else:
        # По умолчанию пробуем как CSV
        return _parse_csv_preview(content)


def _detect_delimiter(text: str) -> str:
    """
    Определяет разделитель CSV файла.
    
    Проверяет в порядке приоритета: точка с запятой, запятая, табуляция.
    Для русских CSV (из Excel) обычно используется точка с запятой.
    """
    # Берём первые несколько строк для анализа
    sample_lines = text.split('\n')[:5]
    sample = '\n'.join(sample_lines)
    
    # Считаем количество каждого разделителя
    delimiters = [';', ',', '\t']
    counts = {}
    
    for delim in delimiters:
        # Считаем сколько раз разделитель встречается в каждой строке
        line_counts = [line.count(delim) for line in sample_lines if line.strip()]
        if line_counts:
            # Если разделитель встречается одинаковое количество раз в каждой строке - это хороший признак
            if len(set(line_counts)) == 1 and line_counts[0] > 0:
                counts[delim] = line_counts[0] * 10  # Бонус за консистентность
            else:
                counts[delim] = sum(line_counts)
    
    # Если точка с запятой встречается чаще или так же часто как запятая - используем её
    # (русские CSV из Excel обычно используют точку с запятой)
    detected_delimiter: str
    if counts.get(';', 0) >= counts.get(',', 0) and counts.get(';', 0) > 0:
        detected_delimiter = ';'
    elif counts.get(',', 0) > 0:
        detected_delimiter = ','
    elif counts.get('\t', 0) > 0:
        detected_delimiter = '\t'
    else:
        # Fallback - пробуем sniffer
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            detected_delimiter = dialect.delimiter
        except csv.Error:
            detected_delimiter = ','  # По умолчанию запятая
    
    logger.debug(
        "csv_delimiter_detected",
        delimiter=repr(detected_delimiter),
        counts=counts,
        sample_preview=sample[:100],
    )
    
    return detected_delimiter


def _decode_csv_content(content: bytes) -> str:
    """Декодирует содержимое CSV файла с определением кодировки."""
    # Пробуем определить кодировку
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1251")  # Русская Windows кодировка
        except UnicodeDecodeError:
            text = content.decode("latin-1")
    return text


def _parse_csv_preview(
    content: bytes,
) -> tuple[list[str], int, list[list[Any]]]:
    """Парсит CSV файл."""
    text = _decode_csv_content(content)
    
    # Определяем разделитель
    delimiter = _detect_delimiter(text)
    
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    
    # Фильтруем пустые строки
    rows = [row for row in rows if any(cell.strip() for cell in row)]
    
    if not rows:
        return [], 0, []
    
    columns = rows[0]
    data_rows = rows[1:]
    
    return columns, len(data_rows), data_rows[:PREVIEW_ROWS]


def _parse_excel_preview(
    content: bytes,
) -> tuple[list[str], int, list[list[Any]]]:
    """Парсит Excel файл."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel support requires openpyxl library")
    
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return [], 0, []
    
    columns = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = [[c for c in row] for row in rows[1:]]
    
    return columns, len(data_rows), data_rows[:PREVIEW_ROWS]


def suggest_column_mapping(
    file_columns: list[str],
    directory_columns: list[dict],
) -> dict[str, str | None]:
    """
    Предлагает маппинг колонок файла на колонки справочника.
    
    Использует нечёткое сравнение названий.
    """
    mapping: dict[str, str | None] = {}
    
    dir_cols = {col["name"]: col["label"] for col in directory_columns}
    dir_labels_lower = {col["label"].lower(): col["name"] for col in directory_columns}
    dir_names_lower = {col["name"].lower(): col["name"] for col in directory_columns}
    
    for file_col in file_columns:
        file_col_lower = file_col.lower().strip()
        
        # Точное совпадение по label
        if file_col_lower in dir_labels_lower:
            mapping[file_col] = dir_labels_lower[file_col_lower]
            continue
        
        # Точное совпадение по name
        if file_col_lower in dir_names_lower:
            mapping[file_col] = dir_names_lower[file_col_lower]
            continue
        
        # Частичное совпадение
        matched = False
        for label_lower, name in dir_labels_lower.items():
            if label_lower in file_col_lower or file_col_lower in label_lower:
                mapping[file_col] = name
                matched = True
                break
        
        if not matched:
            mapping[file_col] = None
    
    return mapping


async def import_file_data(
    db: AsyncSession,
    file: Any,
    mapping: dict[str, str | None],
    has_header: bool,
    directory: Directory,
    tenant_id: Any,
    max_items: int,
    validate_fn: Callable[[dict, int], list[str]],
) -> tuple[int, int, list[dict[str, Any]]]:
    """
    Импортирует данные из файла в справочник.
    
    Args:
        db: Сессия БД
        file: UploadFile объект
        mapping: Маппинг колонок файла → колонки справочника
        has_header: Есть ли заголовок
        directory: Справочник
        tenant_id: ID тенанта
        max_items: Максимальное количество записей для импорта
        validate_fn: Функция валидации данных
    
    Returns:
        Tuple из:
        - created: количество созданных записей
        - skipped: количество пропущенных записей
        - errors: список ошибок
    """
    content = await file.read()
    
    if len(content) > MAX_FILE_SIZE:
        raise ValueError(f"File too large. Maximum size is {MAX_FILE_SIZE // 1024 // 1024} MB")
    
    filename = getattr(file, "filename", "") or ""
    
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        rows = _parse_excel_rows(content, has_header)
        file_columns = _get_excel_columns(content) if has_header else None
    else:
        rows = _parse_csv_rows(content, has_header)
        file_columns = _get_csv_columns(content) if has_header else None
    
    # Если нет заголовка, используем индексы как ключи
    if not has_header or file_columns is None:
        file_columns = [str(i) for i in range(len(rows[0]) if rows else 0)]
    
    # Создаём обратный маппинг: индекс колонки файла → имя колонки справочника
    col_mapping: dict[int, str] = {}
    for i, file_col in enumerate(file_columns):
        dir_col = mapping.get(file_col)
        if dir_col:
            col_mapping[i] = dir_col
    
    created = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    
    for row_num, row in enumerate(rows, start=2 if has_header else 1):
        if created >= max_items:
            skipped += 1
            continue
        
        # Формируем данные записи
        data: dict[str, Any] = {}
        for col_idx, value in enumerate(row):
            if col_idx in col_mapping:
                dir_col_name = col_mapping[col_idx]
                # Пропускаем пустые значения
                if value is not None and value != "":
                    data[dir_col_name] = value
        
        # Валидируем
        validation_errors = validate_fn(data, row_num)
        if validation_errors:
            errors.append({"row": row_num, "error": "; ".join(validation_errors)})
            continue
        
        # Создаём запись
        item = DirectoryItem(
            tenant_id=tenant_id,
            directory_id=directory.id,
            data=data,
        )
        db.add(item)
        created += 1
    
    logger.info(
        "file_data_imported",
        directory_id=str(directory.id),
        created=created,
        skipped=skipped,
        errors=len(errors),
    )
    
    return created, skipped, errors


def _parse_csv_rows(content: bytes, has_header: bool) -> list[list[Any]]:
    """Парсит строки CSV."""
    text = _decode_csv_content(content)
    
    # Определяем разделитель
    delimiter = _detect_delimiter(text)
    
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    
    # Фильтруем пустые строки
    rows = [row for row in rows if any(cell.strip() for cell in row)]
    
    if has_header and rows:
        rows = rows[1:]  # Пропускаем заголовок
    
    return rows


def _get_csv_columns(content: bytes) -> list[str]:
    """Получает колонки CSV."""
    text = _decode_csv_content(content)
    
    # Определяем разделитель
    delimiter = _detect_delimiter(text)
    
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    
    # Ищем первую непустую строку
    for row in rows:
        if any(cell.strip() for cell in row):
            return row
    
    return []


def _parse_excel_rows(content: bytes, has_header: bool) -> list[list[Any]]:
    """Парсит строки Excel."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel support requires openpyxl library")
    
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    
    if has_header and rows:
        rows = rows[1:]
    
    return rows


def _get_excel_columns(content: bytes) -> list[str]:
    """Получает колонки Excel."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel support requires openpyxl library")
    
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True, max_row=1))
    wb.close()
    
    if rows:
        return [str(c) if c is not None else "" for c in rows[0]]
    return []
